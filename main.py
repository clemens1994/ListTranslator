import openpyxl
from openpyxl import *

import googletrans
from googletrans import *


#---------------------------------------------------------

path       = None
excel_file = None
sheet      = None
translator = googletrans.Translator()


#---------------------------------------------------------

def limit_length(message):
    if len(message) == 0:
        return
    elif len(message) > 73:
        return message[0:73]
    else:
        return message


#---------------------------------------------------------

path = input()

excel_file = openpyxl.load_workbook( path )

sheet = excel_file[excel_file.sheetnames[0]]


#---------------------------------------------------------

for generator in sheet.iter_cols(min_col=4, max_col=4, min_row=2):
    for cell in generator:
            try:
                #translate old cell
                text = str(cell.value)
                translated = translator.translate(text, dest='sr', src='de')

                #add new cell
                new_cell = sheet.cell( row=cell.row, column=cell.column+1)
                new_cell.value = limit_length( str(translated.text) )
                print(new_cell.value)
            except:
                #add new cell (blank cell)
                new_cell = sheet.cell( row=cell.row, column=cell.column+1)
                new_cell.value = empty_string = ''


#---------------------------------------------------------

excel_file.save('t100_uebersetzt')
